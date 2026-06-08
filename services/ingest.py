"""Extract plain text from uploaded source files for grounded generation.

Supported types: .txt, .md, .csv, .xlsx, .pdf, .docx. The extracted text is
truncated to MAX_CHARS to keep prompt token usage bounded for the prototype.
"""

from __future__ import annotations

import csv
import io

MAX_CHARS = 12_000

SUPPORTED_EXTENSIONS = ("txt", "md", "csv", "xlsx", "pdf", "docx")


class UnsupportedFileType(ValueError):
    pass


def extract_text(filename: str, data: bytes) -> str:
    """Return plain text extracted from file bytes, dispatched by extension.

    Args:
        filename: original file name (used only for its extension).
        data: raw file bytes.
    """
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    if ext in ("txt", "md"):
        text = _from_plain(data)
    elif ext == "csv":
        text = _from_csv(data)
    elif ext == "xlsx":
        text = _from_xlsx(data)
    elif ext == "pdf":
        text = _from_pdf(data)
    elif ext == "docx":
        text = _from_docx(data)
    else:
        raise UnsupportedFileType(
            f"Unsupported file type: .{ext or '?'}. "
            f"Supported: {', '.join('.' + e for e in SUPPORTED_EXTENSIONS)}"
        )

    text = text.strip()
    if len(text) > MAX_CHARS:
        text = text[:MAX_CHARS].rstrip() + "\n\n[...truncated for length...]"
    return text


def _from_plain(data: bytes) -> str:
    return data.decode("utf-8", errors="replace")


def _from_csv(data: bytes) -> str:
    buffer = io.StringIO(data.decode("utf-8-sig", errors="replace"))
    rows = csv.reader(buffer)
    return "\n".join(", ".join(cell for cell in row) for row in rows)


def _from_xlsx(data: bytes) -> str:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    lines: list[str] = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(values_only=True):
            cells = [str(cell) for cell in row if cell is not None]
            if cells:
                lines.append(", ".join(cells))
    return "\n".join(lines)


def _from_pdf(data: bytes) -> str:
    from pypdf import PdfReader

    reader = PdfReader(io.BytesIO(data))
    pages = [page.extract_text() or "" for page in reader.pages]
    return "\n\n".join(pages)


def _from_docx(data: bytes) -> str:
    from docx import Document

    document = Document(io.BytesIO(data))
    return "\n".join(p.text for p in document.paragraphs if p.text)

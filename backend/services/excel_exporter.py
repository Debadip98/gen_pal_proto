"""GenPal Excel export and post-export validation."""

from __future__ import annotations

import io

from backend.core import constants


def order_and_title(rows: list[dict]) -> list[dict]:
    """Sort by career-level order then complexity order, then assign title 1..N."""
    ordered = sorted(
        rows,
        key=lambda r: (
            constants.level_sort_key(r.get("career_level", "")),
            constants.complexity_sort_key(r.get("complexity", "")),
        ),
    )
    for index, row in enumerate(ordered, start=1):
        row["title"] = index
    return ordered


def to_xlsx_bytes(rows: list[dict]) -> bytes:
    """Write rows to a single ``Sheet1`` with exactly the 11 GenPal columns."""
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = constants.EXCEL_SHEET_NAME
    sheet.append(constants.GENPAL_COLUMNS)
    for row in rows:
        sheet.append([_cell(row, column) for column in constants.GENPAL_COLUMNS])

    widths = (8, 14, 22, 26, 14, 12, 14, 60, 60, 10, 45)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _cell(row: dict, column: str):
    value = row.get(column, "")
    if column == "options":
        return None
    return value


def validate_workbook(
    xlsx_bytes: bytes, ssid: str, expected_count: int
) -> list[str]:
    """Reopen the saved workbook and verify it matches the contract."""
    from openpyxl import load_workbook

    errors: list[str] = []
    workbook = load_workbook(io.BytesIO(xlsx_bytes))

    if len(workbook.sheetnames) != 1:
        errors.append(f"Workbook must have exactly 1 sheet, found {len(workbook.sheetnames)}.")
    if workbook.sheetnames and workbook.sheetnames[0] != constants.EXCEL_SHEET_NAME:
        errors.append(f"Sheet must be named {constants.EXCEL_SHEET_NAME!r}.")

    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        errors.append("Workbook is empty.")
        return errors

    header = list(rows[0])
    if len(header) != constants.EXCEL_COLUMN_COUNT:
        errors.append(f"Header must have {constants.EXCEL_COLUMN_COUNT} columns, found {len(header)}.")
    if header != constants.GENPAL_COLUMNS:
        errors.append(f"Header order is wrong. Expected {constants.GENPAL_COLUMNS}, got {header}.")

    data = rows[1:]
    if len(data) != expected_count:
        errors.append(f"Expected {expected_count} data rows, found {len(data)}.")

    title_idx = constants.GENPAL_COLUMNS.index("title")
    ssid_idx = constants.GENPAL_COLUMNS.index("ssid")
    qtype_idx = constants.GENPAL_COLUMNS.index("question_type")
    options_idx = constants.GENPAL_COLUMNS.index("options")
    ssid = (ssid or "").strip()

    for offset, data_row in enumerate(data, start=1):
        if data_row[title_idx] != offset:
            errors.append(f"Data row {offset}: title must be {offset}, found {data_row[title_idx]!r}.")
        if str(data_row[ssid_idx]) != ssid:
            errors.append(f"Data row {offset}: ssid must equal {ssid!r}.")
        if data_row[qtype_idx] != constants.QUESTION_TYPE:
            errors.append(f"Data row {offset}: question_type must be {constants.QUESTION_TYPE}.")
        if data_row[options_idx] not in (None, ""):
            errors.append(f"Data row {offset}: options must be blank.")

    return errors

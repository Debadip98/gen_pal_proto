"""Tests for Excel output contract: exactly 11 columns, Sheet1 only."""

from __future__ import annotations

import io

import openpyxl

from backend.core.constants import GENPAL_COLUMNS, EXCEL_SHEET_NAME
from backend.services.excel_exporter import order_and_title, to_xlsx_bytes


def _make_rows(n: int = 3) -> list[dict]:
    return [
        {
            "title": "",
            "ssid": "80002591",
            "skill": "Test Skill",
            "topic": f"Topic {i}",
            "question_type": "QnA",
            "career_level": "ASE",
            "complexity": "Basic",
            "question": f"Question {i}?",
            "answer": f"Answer {i}.",
            "options": "",
            "reference_url": "https://example.com",
        }
        for i in range(n)
    ]


def test_column_count_is_11():
    rows = order_and_title(_make_rows(5))
    xlsx = to_xlsx_bytes(rows)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx))
    ws = wb.active
    assert ws.max_column == 11, f"Expected 11 columns, got {ws.max_column}"


def test_column_names_exact_order():
    rows = order_and_title(_make_rows(2))
    xlsx = to_xlsx_bytes(rows)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx))
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, 12)]
    assert headers == GENPAL_COLUMNS


def test_only_one_sheet():
    rows = order_and_title(_make_rows(2))
    xlsx = to_xlsx_bytes(rows)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx))
    assert wb.sheetnames == [EXCEL_SHEET_NAME]


def test_titles_sequential():
    """Titles are integers 1..N (the GenPal output contract)."""
    rows = order_and_title(_make_rows(5))
    xlsx = to_xlsx_bytes(rows)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx))
    ws = wb.active
    title_col = GENPAL_COLUMNS.index("title") + 1
    titles = [ws.cell(r, title_col).value for r in range(2, ws.max_row + 1)]
    expected = list(range(1, len(titles) + 1))
    assert titles == expected


def test_no_extra_columns():
    rows = order_and_title(_make_rows(1))
    xlsx = to_xlsx_bytes(rows)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx))
    ws = wb.active
    assert ws.cell(1, 12).value is None


def test_row_count_matches():
    n = 7
    rows = order_and_title(_make_rows(n))
    xlsx = to_xlsx_bytes(rows)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx))
    ws = wb.active
    # header row + n data rows
    assert ws.max_row == n + 1
